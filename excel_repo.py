from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


class ExcelRepository:
    CONTRACTS_HEADERS = [
        "Contract No",
        "Customer Name",
        "Customer Address",
        "Installation Address",
        "Product Days",
        "Installation Days",
        "Load Capacity (kg)",
        "Speed (m/min)",
        "Motor Brand",
        "Motor Power",
        "Control System Brand",
        "Number of Stops",
        "Number of Floors Served",
        "Electrical Box Brand",
        "Pit Depth (mm)",
        "Overhead (mm)",
        "Steel Frame",
        "Shaft Wall Material",
        "Cabin Door Material",
        "Landing Door Material",
        "Start Date",
        "Product End",
        "Installation End",
        "Source File",
    ]

    TASKS_HEADERS = [
        "Contract No",
        "Customer Name",
        "Task Type",
        "Start Date",
        "End Date",
    ]

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)

    def ensure_workbook(self):
        if self.file_path.exists():
            wb = load_workbook(self.file_path)
        else:
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            wb.create_sheet("Contracts")
            wb.create_sheet("Tasks")

        contracts_ws = wb["Contracts"] if "Contracts" in wb.sheetnames else wb.create_sheet("Contracts")
        tasks_ws = wb["Tasks"] if "Tasks" in wb.sheetnames else wb.create_sheet("Tasks")

        self._ensure_headers(contracts_ws, self.CONTRACTS_HEADERS)
        self._ensure_headers(tasks_ws, self.TASKS_HEADERS)

        wb.save(self.file_path)
        return wb

    def save_contract_and_plan(self, contract, plan):
        wb = self.ensure_workbook()
        contracts_ws = wb["Contracts"]
        tasks_ws = wb["Tasks"]

        contract_row = [
            contract.contract_no,
            contract.customer_name,
            contract.customer_address,
            contract.installation_address,
            contract.product_days,
            contract.installation_days,
            contract.load_capacity_kg,
            contract.speed_mpm,
            contract.motor_brand,
            contract.motor_power,
            contract.control_system_brand,
            contract.number_of_stops,
            contract.number_of_floors_served,
            contract.electrical_box_brand,
            contract.pit_depth_mm,
            contract.overhead_mm,
            "YES" if contract.steel_frame else "NO",
            contract.shaft_wall_material,
            contract.cabin_door_material,
            contract.landing_door_material,
            plan.start_date.date().isoformat(),
            plan.product_end.date().isoformat(),
            plan.installation_end.date().isoformat(),
            contract.source_file,
        ]

        existing_rows = self._find_contract_rows(contracts_ws, contract.contract_no)
        if existing_rows:
            target_row = existing_rows[0]
            self._write_row(contracts_ws, target_row, contract_row)
            for duplicate_row in reversed(existing_rows[1:]):
                contracts_ws.delete_rows(duplicate_row, 1)
        else:
            contracts_ws.append(contract_row)

        self._delete_task_rows(tasks_ws, contract.contract_no)
        tasks_ws.append([
            contract.contract_no,
            contract.customer_name,
            "Product",
            plan.product_start.date().isoformat(),
            plan.product_end.date().isoformat(),
        ])
        tasks_ws.append([
            contract.contract_no,
            contract.customer_name,
            "Installation",
            plan.installation_start.date().isoformat(),
            plan.installation_end.date().isoformat(),
        ])

        wb.save(self.file_path)

    def _ensure_headers(self, ws, headers: Iterable[str]) -> None:
        headers = list(headers)
        current = [ws.cell(row=1, column=i + 1).value for i in range(len(headers))]
        if current != headers:
            if ws.max_row == 0:
                ws.append(headers)
                return
            for i, header in enumerate(headers, start=1):
                ws.cell(row=1, column=i).value = header

    def _find_contract_rows(self, ws, contract_no: str) -> list[int]:
        if not contract_no:
            return []
        normalized = contract_no.strip().lower()
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=1).value
            if value is None:
                continue
            if str(value).strip().lower() == normalized:
                rows.append(row_idx)
        return rows

    def _write_row(self, ws, row_idx: int, values: list) -> None:
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    def _delete_task_rows(self, ws, contract_no: str) -> None:
        if not contract_no:
            return
        normalized = contract_no.strip().lower()
        rows_to_delete = []
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=1).value
            if value is None:
                continue
            if str(value).strip().lower() == normalized:
                rows_to_delete.append(row_idx)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx, 1)
